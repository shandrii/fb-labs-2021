from openpyxl import Workbook
import math

a = ord('а')
path = 'TEXT.txt'


def filter_with_spaces(path):
    input_file = open(path, 'r', encoding='cp1251')
    input_text = input_file.read()
    text = ''
    input_file.close()

    for letter in input_text.lower():
        if letter in ' абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
            text += letter
    text = ' '.join(text.split())
    with_space = open('TEXT_with_spaces.txt', 'w', encoding='utf-8')
    with_space.write(text)
    with_space.close()


def filter_without_spaces(path):
    input_file = open(path, 'r', encoding='cp1251')
    input_text = input_file.read()
    text = ''
    input_file.close()
    for letter in input_text.lower():
        if letter in 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
            text += letter
    without_space = open('TEXT_without_spaces.txt', 'w', encoding='utf-8')
    without_space.write(text)
    without_space.close()


def count_frequency(path):
    # 'TEXT_without_spaces.txt'
    input_text = open(path, 'r', encoding='utf-8')
    text = input_text.read()
    input_text.close()
    length = len(text)
    frequency = [text.count(chr(i)) / length for i in range(a, a + 32)]
    if text.find(' ') != -1:
        frequency.append(text.count(' ') / length)
    return frequency


def count_bigram_frequency(path):
    input_text = open(path, 'r', encoding='utf-8')
    text = input_text.read()
    input_text.close()
    length = len(text)
    bigrams_count = {}
    bigrams_frequency = {}
    for i in ' абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
        for j in ' абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
            bigrams_count[i + j] = 0
    for i in range(0, length - 1):
        bigram = text[i:i + 2]
        bigrams_count[bigram] += 1
    for key in bigrams_count.keys():
        bigrams_frequency[key] = bigrams_count[key] / length
    return bigrams_frequency


def count_bigram_frequency2(path):
    input_text = open(path, 'r', encoding='utf-8')
    text = input_text.read()
    input_text.close()
    length = len(text)
    bigrams_count = {}
    bigrams_frequency = {}
    for i in ' абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
        for j in ' абвгдеёжзийклмнопрстуфхцчшщъыьэюя':
            bigrams_count[i + j] = 0
    if length % 2:
        text = text[0:-1]
        length -= 1
    for i in range(0, length - 1, 2):
        bigram = text[i:i + 2]
        bigrams_count[bigram] += 1
    for key in bigrams_count.keys():
        bigrams_frequency[key] = bigrams_count[key] / (length / 2)
    return bigrams_frequency


def filling(arr, flag):
    wb = Workbook()
    ws = wb.active
    for i in range(32):
        ws.cell(i + 1, 1).value = chr(a + i)
        ws.cell(i + 1, 2).value = arr[i]
    if flag:
        ws.cell(33, 1).value = '\'  \''
        ws.cell(33, 2).value = arr[-1]
        wb.save('results_for_letters_with_space.xlsx')
    else:
        wb.save('results_for_letters_without_space.xlsx')


def filling_bigrams(bigrams_frequency: dict, withspaces, onestep):
    wb = Workbook()
    ws = wb.active
    k = 1
    i = 1
    for key in bigrams_frequency.keys():
        if k == 129:
            i += 3
            k = 1
        ws.cell(k, i).value = key
        ws.cell(k, i + 1).value = bigrams_frequency.get(key)
        k += 1
    if withspaces:
        if onestep:
            wb.save('results_for_bigrams_with_spaces_onestep.xlsx')
        else:
            wb.save('results_for_bigrams_with_spaces_twosteps.xlsx')
    else:
        if onestep:
            wb.save('results_for_bigrams_without_spaces_onestep.xlsx')
        else:
            wb.save('results_for_bigrams_without_spaces_twosteps.xlsx')


def entropy(frequency_arr):
    h = 0
    for p in frequency_arr:
        h += p * math.log(p, 2)
    return round(-h, 5)


def bigrams_entropy(frequency_dict: dict):
    h = 0
    for p in frequency_dict.values():
        if p == 0:
            continue
        h += p * math.log(p, 2)
    return round(-h / 2, 5)


def redundancy(entropy, withspaces):
    if withspaces:
        return round(1 - (entropy / math.log(33, 2)), 5)
    else:
        return round(1 - (entropy / math.log(32, 2)), 5)


filter_without_spaces(path)
filter_with_spaces(path)
filling(count_frequency('TEXT_with_spaces.txt'), True)
filling(count_frequency('TEXT_without_spaces.txt'), False)
filling_bigrams(count_bigram_frequency('TEXT_with_spaces.txt'), True, True)  # with spaces, one step
filling_bigrams(count_bigram_frequency2('TEXT_with_spaces.txt'), True, False)  # with spaces two steps
filling_bigrams(count_bigram_frequency('TEXT_without_spaces.txt'), False, True)  # without spaces one step
filling_bigrams(count_bigram_frequency2('TEXT_without_spaces.txt'), False, False)  # without spaces two steps
print('Entropy without spaces: ', entropy(count_frequency('TEXT_without_spaces.txt')))
print('Redundancy without spaces: ',
      redundancy(entropy(count_frequency('TEXT_without_spaces.txt', )), False))
print('Entropy with spaces: ', entropy(count_frequency('TEXT_with_spaces.txt')))
print('Redundancy with spaces: ',
      redundancy(entropy(count_frequency('TEXT_with_spaces.txt', )), True))

print('Entropy for bigrams with spaces, one step: ',
      bigrams_entropy(count_bigram_frequency('TEXT_with_spaces.txt')))
print('Redundancy for bigrams with spaces, one step: ',
      redundancy(bigrams_entropy(count_bigram_frequency('TEXT_with_spaces.txt', )), True))
print('Entropy for bigrams without spaces, one step: ',
      bigrams_entropy(count_bigram_frequency('TEXT_without_spaces.txt')))
print('Redundancy for bigrams without spaces, one step: ',
      redundancy(bigrams_entropy(count_bigram_frequency('TEXT_without_spaces.txt', )), False))
print('Entropy for bigrams without spaces, two steps: ',
      bigrams_entropy(count_bigram_frequency2('TEXT_without_spaces.txt')))
print('Redundancy for bigrams without spaces, two steps: ',
      redundancy(bigrams_entropy(count_bigram_frequency2('TEXT_without_spaces.txt')), False))
print('Entropy for bigrams with spaces, two steps: ',
      bigrams_entropy(count_bigram_frequency2('TEXT_with_spaces.txt')))
print('Redundancy for bigrams with spaces, two steps: ',
      redundancy(bigrams_entropy(count_bigram_frequency2('TEXT_with_spaces.txt')), True))
